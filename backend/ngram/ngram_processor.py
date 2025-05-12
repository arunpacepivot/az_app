import pandas as pd
from sklearn.feature_extraction.text import CountVectorizer
from collections import defaultdict
from typing import Dict, List
import warnings
import os
from openpyxl.utils.dataframe import dataframe_to_rows
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def load_data(file_path: str, sheet_name: str) -> pd.DataFrame:
    """Load data from an Excel file."""
    return pd.read_excel(file_path, sheet_name=sheet_name)

def extract_asin(data: pd.DataFrame) -> pd.DataFrame:
    data["ASIN"] = data["Campaign Name (Informational only)"].apply(lambda x: x.split()[0] if isinstance(x, str) else None)
    return data

def summarize_data(data: pd.DataFrame) -> pd.DataFrame:
    data_summary = data.groupby("ASIN").agg({
        "Impressions": "sum",
        "Clicks": "sum",
        "Spend": "sum",
        "Sales": "sum",
        "Orders": "sum",
        "Units": "sum"
    }).reset_index()
    
    data_summary["CPC"] = data_summary["Spend"] / data_summary["Clicks"]
    data_summary["RPC"] = data_summary["Sales"] / data_summary["Clicks"]
    data_summary["AOV"] = data_summary["Sales"] / data_summary["Units"]
    data_summary["Conversion"] = data_summary["Orders"] / data_summary["Clicks"]
    return data_summary

def filter_data_sk(data: pd.DataFrame) -> pd.DataFrame:
    """Filter rows based on ASIN starting with 'B0'."""
    filtered_data = data[data['ASIN'].str.startswith('B0', na=False)]
    filtered_data = filtered_data[~filtered_data['Customer Search Term'].str.startswith('b0', na=False)]
    return filtered_data

def filter_data_mk(data: pd.DataFrame) -> pd.DataFrame:
    """Filter rows based on ASIN not starting with 'B0'."""
    filtered_data_mk = data[~data['ASIN'].str.startswith('B0', na=False)]
    filtered_data_mk = filtered_data_mk[~filtered_data_mk['Customer Search Term'].str.startswith('b0', na=False)]
    return filtered_data_mk

def update_ngram_metrics(
    terms: pd.Series, 
    ngram_range: tuple, 
    metric_data: pd.DataFrame, 
    ngram_type: str, 
    asin: str, 
    asin_ngram_metrics: Dict[str, Dict[str, Dict[str, Dict[str, float]]]],
    metrics: List[str]
) -> None:
    valid_mask = terms.str.strip().str.len() > 0
    if not valid_mask.any():
        print(f"Warning: No valid terms found for ASIN {asin} with {ngram_type}")
        return

    terms_filtered = terms[valid_mask]
    metric_data_filtered = metric_data[valid_mask]
    vectorizer = CountVectorizer(
        ngram_range=ngram_range,
        stop_words=None,
        token_pattern=r"(?u)\b\w+\b",
        min_df=1
    )

    try:
        X = vectorizer.fit_transform(terms_filtered)
        ngram_terms = vectorizer.get_feature_names_out()
        
        for row_idx, ngram_counts in enumerate(X):
            row_data = metric_data_filtered.iloc[row_idx]
            for idx, ngram in enumerate(ngram_terms):
                count = ngram_counts[0, idx]
                if count > 0:
                    for metric in metrics:
                        asin_ngram_metrics[asin][ngram_type][ngram][metric] += row_data[metric] * count
    except ValueError as e:
        print(f"Warning: Could not process {ngram_type} for ASIN {asin}: {str(e)}")

def perform_ngram_analysis(data_filtered: pd.DataFrame, metrics: List[str]) -> Dict[str, Dict[str, Dict[str, Dict[str, float]]]]:
    """Perform n-gram analysis on filtered data."""
    asin_ngram_metrics = defaultdict(lambda: {
        'unigram': defaultdict(lambda: {metric: 0 for metric in metrics}),
        'bigram': defaultdict(lambda: {metric: 0 for metric in metrics}),
        'trigram': defaultdict(lambda: {metric: 0 for metric in metrics}),
    })

    for asin, group in data_filtered.groupby('ASIN'):
        update_ngram_metrics(group['Customer Search Term'], (1, 1), group[metrics], 'unigram', asin, asin_ngram_metrics, metrics)
        update_ngram_metrics(group['Customer Search Term'], (2, 2), group[metrics], 'bigram', asin, asin_ngram_metrics, metrics)
        update_ngram_metrics(group['Customer Search Term'], (3, 3), group[metrics], 'trigram', asin, asin_ngram_metrics, metrics)
    
    return asin_ngram_metrics

def save_ngram_analysis(
    asin_ngram_metrics: Dict[str, Dict[str, Dict[str, Dict[str, float]]]], 
    data_summary: pd.DataFrame, 
    output_path: str,
    bulk_data: pd.DataFrame,
    target_acos: float
) -> None:
    """Save n-gram analysis results to an Excel file."""
    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        for asin, ngram_data in asin_ngram_metrics.items():
            unigram_df = pd.DataFrame.from_dict(ngram_data['unigram'], orient='index')
            unigram_df['RPC'] = unigram_df.apply(lambda row: row['Sales'] / row['Clicks'] if row['Clicks'] > 0 else 0, axis=1)
            unigram_df['ACOS'] = unigram_df.apply(lambda row: row['Spend'] / row['Sales'] if row['Sales'] > 0 else 0, axis=1)
            
            asin_data_summary = data_summary[data_summary['ASIN'] == asin]
            if not asin_data_summary.empty:
                aov = asin_data_summary['AOV'].values[0]
                filtered_unigram_df = unigram_df[(unigram_df['Orders'] == 0) | (unigram_df['ACOS'] > target_acos * 1.2)]
                filtered_unigram_df = filtered_unigram_df[filtered_unigram_df['Spend'] > aov * target_acos]
            else:
                filtered_unigram_df = unigram_df

            filtered_unigram_df.to_excel(writer, sheet_name=f"{asin}", startrow=0, startcol=0)

            # Add check for empty bigram DataFrame
            bigram_df = pd.DataFrame.from_dict(ngram_data['bigram'], orient='index')
            if not bigram_df.empty:
                bigram_df['RPC'] = bigram_df.apply(lambda row: row['Sales'] / row['Clicks'] if row['Clicks'] > 0 else 0, axis=1)
                bigram_df['ACOS'] = bigram_df.apply(lambda row: row['Spend'] / row['Sales'] if row['Sales'] > 0 else 0, axis=1)
                
                asin_data_summary = data_summary[data_summary['ASIN'] == asin]
                if not asin_data_summary.empty:
                    aov = asin_data_summary['AOV'].values[0]
                    filtered_bigram_df = bigram_df[(bigram_df['Orders'] == 0) | (bigram_df['ACOS'] > target_acos * 1.2)]
                    filtered_bigram_df = filtered_bigram_df[filtered_bigram_df['Spend'] > aov * target_acos]
                else:
                    filtered_bigram_df = bigram_df

                filtered_bigram_df.to_excel(writer, sheet_name=f"{asin}", startrow=0, startcol=filtered_unigram_df.shape[1] + 3)
            else:
                filtered_bigram_df = pd.DataFrame()  # Empty DataFrame for consistent reference

            # Similar check for trigram DataFrame
            trigram_df = pd.DataFrame.from_dict(ngram_data['trigram'], orient='index')
            if not trigram_df.empty:
                trigram_df['RPC'] = trigram_df.apply(lambda row: row['Sales'] / row['Clicks'] if row['Clicks'] > 0 else 0, axis=1)
                trigram_df['ACOS'] = trigram_df.apply(lambda row: row['Spend'] / row['Sales'] if row['Sales'] > 0 else 0, axis=1)
                
                asin_data_summary = data_summary[data_summary['ASIN'] == asin]
                if not asin_data_summary.empty:
                    aov = asin_data_summary['AOV'].values[0]
                    filtered_trigram_df = trigram_df[(trigram_df['Orders'] == 0) | (trigram_df['ACOS'] > target_acos * 1.2)]
                    filtered_trigram_df = filtered_trigram_df[filtered_trigram_df['Spend'] > aov * target_acos]
                else:
                    filtered_trigram_df = trigram_df

                filtered_trigram_df.to_excel(writer, sheet_name=f"{asin}", startrow=0, startcol=filtered_unigram_df.shape[1] + filtered_bigram_df.shape[1] + 6)
            else:
                filtered_trigram_df = pd.DataFrame()  # Empty DataFrame for consistent reference
    
    # Create a new DataFrame from bulk_data containing unique Ad Group IDs and their respective Campaign Name, Campaign ID, and Ad Group ID
    unique_ad_groups_df = bulk_data[["Ad Group ID", "Campaign Name (Informational only)", "Campaign ID", "Ad Group Name (Informational only)"]].drop_duplicates(subset=["Ad Group ID"])
    unique_ad_groups_df.reset_index(drop=True, inplace=True)
    
    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a") as writer:
        for asin in asin_ngram_metrics.keys():
            # Filter campaigns that start with the ASIN
            asin_campaigns = unique_ad_groups_df[unique_ad_groups_df["Campaign Name (Informational only)"].str.startswith(asin, na=False)]
            
            # Further filter for campaigns containing 'broad' or 'auto' in the campaign name
            filtered_campaigns = asin_campaigns[
                asin_campaigns["Campaign Name (Informational only)"].str.contains("broad", case=False, na=False) |
                asin_campaigns["Campaign Name (Informational only)"].str.contains("auto", case=False, na=False)
            ]
            
            # Select relevant columns
            campaign_df = filtered_campaigns[["Campaign ID", "Ad Group ID", "Campaign Name (Informational only)", "Ad Group Name (Informational only)"]].copy()
            
            # Load the existing workbook
            book = writer.book
            if asin in book.sheetnames:
                sheet = book[asin]
                
                # Find the lowest row of the existing data
                max_row = max(sheet.max_row for sheet in [sheet])
                
                # Write the new DataFrame to the sheet after leaving 3 blank rows
                start_row = max_row + 3
                for r_idx, row in enumerate(dataframe_to_rows(campaign_df, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)
            else:
                # If the sheet does not exist, create it and write the DataFrame
                campaign_df.to_excel(writer, sheet_name=asin, startrow=0, index=False)

def process_ngram_file(bulk_file_path, output_path_sk, output_path_mk, target_acos=0.2):
    """Main processing function for ngram analysis."""
    try:
        # Load data
        data = load_data(bulk_file_path, 'SP Search Term Report')
        bulk_data = load_data(bulk_file_path, 'Sponsored Products Campaigns')

        # Preprocess and summarize data
        data = extract_asin(data)
        data_summary = summarize_data(data)

        # Filter data and define metrics
        filtered_data_sk = filter_data_sk(data)
        filtered_data_mk = filter_data_mk(data)
        metrics = ['Impressions', 'Clicks', 'Spend', 'Sales', 'Orders', 'Units']
        
        # Perform n-gram analysis
        asin_ngram_metrics_sk = perform_ngram_analysis(filtered_data_sk, metrics)
        asin_ngram_metrics_mk = perform_ngram_analysis(filtered_data_mk, metrics)
        
        # Save results
        save_ngram_analysis(asin_ngram_metrics_sk, data_summary, output_path_sk, bulk_data, target_acos)
        save_ngram_analysis(asin_ngram_metrics_mk, data_summary, output_path_mk, bulk_data, target_acos)
        
        return {
            "status": "success",
            "message": "N-gram analysis completed successfully",
            "sk_asin_count": len(asin_ngram_metrics_sk),
            "mk_asin_count": len(asin_ngram_metrics_mk)
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        } 